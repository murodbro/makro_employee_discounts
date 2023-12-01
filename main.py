import sys
import openpyxl
import warnings
import asyncio

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import (
        QApplication, 
        QTableWidgetItem, 
        QTableWidget,
        QWidget, 
        QPushButton,
        QHBoxLayout,
        QVBoxLayout,
        QLabel,
    )

from methods import show_popup, open_file, clean_phone_numbers
from client_requests import add_employees, remove_employees


warnings.filterwarnings("ignore", category=DeprecationWarning)

class MainUiClass(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):

        self.button = QPushButton("Выберите файл", self)
        self.button.setFixedSize(QtCore.QSize(310, 70))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.button.setFont(font)

        self.label = QLabel("Пожалуйста, выберите файл, нажав кнопку", self)
        font = QtGui.QFont()
        font.setPointSize(14)
        self.label.setFont(font)
        self.label.setMaximumHeight(50)

        # Set window properties
        self.setWindowTitle("Обновление сегмента скидок")
        self.setGeometry(0, 0, 1920, 1080)
        self.showMaximized()

        choose_file_layout = QVBoxLayout(self)
        top_layout = QHBoxLayout()
        table_layout = QHBoxLayout()

        choose_file_layout.addStretch(1)
        choose_file_layout.addWidget(self.label, alignment=QtCore.Qt.AlignCenter)
        choose_file_layout.addWidget(self.button, alignment=QtCore.Qt.AlignCenter)
        choose_file_layout.addStretch(1)

        # Excel rider
        choose_file_layout.addLayout(top_layout)
        choose_file_layout.addLayout(table_layout)

        self.update_button = QPushButton("Обновлять", self)
        self.update_button.setMaximumHeight(25)
        font = QtGui.QFont()
        font.setPointSize(8)
        self.update_button.setFont(font)
        self.update_button.setEnabled(False)
        self.update_button.hide()

        self.open_button = QPushButton("Открыть другой файл", self)
        self.open_button.setMaximumHeight(25)
        font = QtGui.QFont()
        font.setPointSize(8)
        self.open_button.setFont(font)
        self.open_button.hide()

        self.lineEdit = QtWidgets.QLineEdit(self)
        self.lineEdit.setFixedSize(QtCore.QSize(200, 25))
        font = QtGui.QFont()
        font.setPointSize(8)
        self.lineEdit.setFont(font)
        self.lineEdit.setPlaceholderText("Сегмент ID")
        self.lineEdit.textChanged.connect(self.toggle_update_button)
        self.lineEdit.hide()

        top_layout.addStretch(1)
        top_layout.addWidget(self.lineEdit, alignment=QtCore.Qt.AlignCenter)
        top_layout.addWidget(self.update_button, alignment=QtCore.Qt.AlignCenter)
        top_layout.addWidget(self.open_button, alignment=QtCore.Qt.AlignRight)

        self.tableWidget = QTableWidget(self)
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.tableWidget.hide()

        table_layout.addWidget(self.tableWidget)

        self.setLayout(choose_file_layout)

        self.button.clicked.connect(self.choose_file)
        self.open_button.clicked.connect(self.open_other_file)
        self.update_button.clicked.connect(lambda: asyncio.run(self.send_request()))
    

    def toggle_update_button(self):
        self.update_button.setEnabled(bool(self.lineEdit.text()))


    async def send_request(self):
        QApplication.setOverrideCursor(QtGui.QCursor(QtCore.Qt.WaitCursor))
        self.update_button.setEnabled(False)

        try:
            self.separate()
        except Exception:
            return show_popup("Ошибка извлечения данных")
        finally:
            QApplication.restoreOverrideCursor()
            self.update_button.setEnabled(True)

        try:
            adding = asyncio.create_task(add_employees(segment_id=self.lineEdit.text(), customer_ids=self.works))
            deleting = asyncio.create_task(remove_employees(segment_id=self.lineEdit.text(), customer_ids=self.fireds))

            await asyncio.gather(adding, deleting)
        except Exception as e:
            return show_popup(str(e))
        finally:
            QApplication.restoreOverrideCursor()
            self.update_button.setEnabled(True)

        show_popup("Система обновлена ​​новыми сотрудниками", success=True)


    def separate(self):
        self.works = []
        self.fireds = []
        phone_numbers = list(clean_phone_numbers(self.numbers))
        
        for column in self.data[0]:
            if column == "Статус":
                index = self.data[0].index(column)
                break

        row = 0
        for row_data in self.data[1:]:
            if row_data[index] == "Работает":
                self.works.append(phone_numbers[row])
            else:
                self.fireds.append(phone_numbers[row])
            row += 1


    def resizeEvent(self, event):
        super().resizeEvent(event)

        if hasattr(self, 'tableWidget'):
            new_size_width = self.size() * 0.97
            new_size_height = self.size() * 0.93
            self.tableWidget.setFixedSize(new_size_width.width(), new_size_height.height())


    def open_excel(self, path):
        file_path = path
        workbook = openpyxl.load_workbook(file_path.split(",")[0])
        sheet = workbook.active

        self.tableWidget.setRowCount(sheet.max_row - 1)
        self.tableWidget.setColumnCount(sheet.max_column - 1)
        self.data = list(sheet.values)
        self.tableWidget.setHorizontalHeaderLabels(self.data[0])

        for phone in self.data[0]:
            if phone == "Номер телефона":
                index = self.data[0].index(phone)
                break

        self.numbers = []
        row = 0
        for row_data in self.data[1:]:
            column = 0
            for column_data in row_data:
                self.tableWidget.setItem(row, column, QTableWidgetItem(str(column_data)))
                if column == index:
                    self.numbers.append(column_data)
                column += 1
            row += 1
        self.tableWidget.resizeColumnsToContents()


    def open_other_file(self):

        path = open_file()

        if not path:
            return
        
        try:
            self.open_excel(path)
        except Exception:
            self.update_button.hide()
            self.open_button.hide()
            self.tableWidget.hide()
            self.lineEdit.hide()
            self.button.show()
            self.label.show()

            return show_popup("Ошибка открытия файла")

        self.tableWidget.resizeColumnsToContents()


    def choose_file(self):
        path = open_file()

        if not path:
            return

        self.button.hide()
        self.label.hide()
        try:
            self.open_excel(path)
        except Exception:
            self.button.show()
            self.label.show()
            return show_popup("Ошибка открытия файла")

        self.update_button.show()
        self.open_button.show()
        self.tableWidget.show()
        self.lineEdit.show()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app_icon = QtGui.QIcon()
    app_icon.addFile('img/logo.png', QtCore.QSize(256,256))
    app.setWindowIcon(app_icon)
    window = MainUiClass()
    window.show()
    sys.exit(app.exec_())
